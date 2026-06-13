#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""One-time migration: video-archive.xlsx -> data/video_pipeline.json.

Reads the volunteer production tracker (Yandex/xlsx master) and emits a
reviewable, repo-native JSON: one record per YouTube video, with a normalised
proof-reading stage, transcription-quality flag, assignees, dates and links.

After this migration the JSON is the source of truth; the spreadsheet is
retired. Re-running is idempotent (deterministic output for the same xlsx).

Usage:
    python scripts/migrate_video_pipeline.py            # writes data/video_pipeline.json
    python scripts/migrate_video_pipeline.py --check    # parse + report, write nothing
"""
import sys
import os
import re
import json
import argparse

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "video-archive.xlsx")
OUT = os.path.join(ROOT, "data", "video_pipeline.json")
APP_DATA = os.path.join(ROOT, "app_data.json")

# Canonical, ordered proof-reading pipeline. `order` drives the dashboard
# progress axis; `key` is the stable machine value.
STAGES = [
    {"key": "queued", "order": 0, "label": "в очереди"},
    {"key": "transcribed", "order": 1, "label": "автотранскрибация"},
    {"key": "review", "order": 2, "label": "сверка"},
    {"key": "read1", "order": 3, "label": "первая читка"},
    {"key": "read2", "order": 4, "label": "вторая читка"},
    {"key": "read3", "order": 5, "label": "третья читка"},
    {"key": "prelayout", "order": 6, "label": "предвёрстка"},
    {"key": "layout", "order": 7, "label": "вёрстка"},
    {"key": "done", "order": 8, "label": "готово"},
]
STAGE_ORDER = {s["key"]: s["order"] for s in STAGES}

# Raw `Статус` strings (lower-cased, whitespace-collapsed) -> canonical stage.
STATUS_MAP = {
    "готов к первой читке": "transcribed",
    "готово к первой читке": "transcribed",
    "готово для первой читки": "transcribed",
    "сверка": "review",
    "сверена первая часть": "review",
    "первая читка": "read1",
    "готово для второй читки": "read1",
    "готово к третьей читке / готово к третьей читке": "read2",
    "третья читка": "read3",
    "готово к предвёрстке": "read3",
    "готово к вёрстке": "prelayout",
    "верстка": "layout",
    "вёрстка": "layout",
}


def collapse(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def yt_id(url):
    m = re.search(r"(?:v=|youtu\.be/)([A-Za-z0-9_-]{6,})", str(url or ""))
    return m.group(1) if m else None


def dur_to_seconds(v):
    """xlsx duration cell -> int seconds. Handles time objects and H:M:S text."""
    if v is None:
        return None
    if hasattr(v, "hour"):  # datetime.time
        return v.hour * 3600 + v.minute * 60 + v.second
    parts = re.findall(r"\d+", str(v))
    if not parts:
        return None
    parts = [int(p) for p in parts]
    while len(parts) < 3:
        parts.insert(0, 0)
    h, m, s = parts[-3], parts[-2], parts[-1]
    return h * 3600 + m * 60 + s


def hms(sec):
    if sec is None:
        return None
    h, rem = divmod(sec, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def classify_transcription(raw):
    """raw `Автотранскрибация` text -> (quality_flag, raw_note)."""
    r = collapse(raw).lower()
    if not r:
        return "unknown", ""
    if r.startswith("готово") or r == "не требуется (по ссылке текст)":
        return "ok", collapse(raw)
    if "без звука" in r or "без каких-то осмысленных" in r:
        return "no_audio", collapse(raw)
    # partial / looped / poor recognition
    return "partial", collapse(raw)


def parse_people_dates(cell):
    """Multiline proof cell -> {assignee, dates[]}. Dates are dd.mm.yy(yy)."""
    txt = str(cell or "").strip()
    if not txt:
        return None
    lines = [ln.strip() for ln in txt.splitlines() if ln.strip()]
    dates = []
    people = []
    for ln in lines:
        if re.fullmatch(r"\d{1,2}\.\d{1,2}\.\d{2,4}", ln):
            dates.append(ln)
        else:
            people.append(ln)
    return {
        "assignee": ", ".join(people) if people else None,
        "dates": dates,
    }


def normalise_date(v):
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month"):  # date/datetime
        return v.strftime("%d.%m.%Y")
    return collapse(v) or None


def infer_stage(status_raw, transcription_flag, proof):
    """Resolve canonical stage from explicit status, else infer."""
    key = STATUS_MAP.get(collapse(status_raw).lower())
    if key:
        return key, False
    # No explicit status: infer from how far the proof columns are filled.
    inferred = None
    if proof["layout"]:
        inferred = "layout"
    elif proof["prelayout"]:
        inferred = "prelayout"
    elif proof["read3"]:
        inferred = "read3"
    elif proof["read2"]:
        inferred = "read2"
    elif proof["read1"]:
        inferred = "read1"
    elif proof["review"]:
        inferred = "review"
    elif transcription_flag in ("ok",):
        inferred = "transcribed"
    else:
        inferred = "queued"
    return inferred, True


def load_catalog_ids():
    try:
        cat = json.load(open(APP_DATA, encoding="utf-8")).get("video_catalog", [])
    except FileNotFoundError:
        return set()
    return {v.get("id") for v in cat if v.get("id")}


def build():
    from openpyxl import load_workbook

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["video_info_new"]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if r and r[0]]

    catalog_ids = load_catalog_ids()
    by_id = {}
    no_id = []
    unmapped_statuses = set()

    for r in data_rows:
        # column indices per the sheet header (see migrate header dump)
        r = tuple(r) + (None,) * (20 - len(r))  # pad ragged rows to full width
        title = collapse(r[0])
        url = r[15]
        vid = yt_id(url)
        dur_sec = dur_to_seconds(r[14])
        tflag, tnote = classify_transcription(r[16])

        proof = {
            "review": parse_people_dates(r[8]),
            "read1": parse_people_dates(r[9]),
            "read2": parse_people_dates(r[10]),
            "read3": parse_people_dates(r[11]),
            "prelayout": parse_people_dates(r[12]),
            "layout": parse_people_dates(r[13]),
        }
        status_raw = collapse(r[2])
        if status_raw and collapse(status_raw).lower() not in STATUS_MAP:
            unmapped_statuses.add(status_raw)
        stage, inferred = infer_stage(status_raw, tflag, proof)

        rec = {
            "id": vid,
            "title": title,
            "purpose": collapse(r[1]) or None,
            "theme": collapse(r[3]) or None,
            "type": collapse(r[4]) or None,
            "priority": collapse(r[5]) or None,
            "duration_sec": dur_sec,
            "duration_hms": hms(dur_sec),
            "youtube_url": collapse(url) or None,
            "status_raw": status_raw or None,
            "stage": stage,
            "stage_inferred": inferred,
            "transcription": {"quality": tflag, "note": tnote or None},
            "assigned": {"issued": normalise_date(r[6]), "submitted": normalise_date(r[7])},
            "proof": {k: v for k, v in proof.items() if v},
            "links": {
                "text": collapse(r[18]) or None,
                "disk_video": collapse(r[19]) or None,
            },
            "notes": collapse(r[17]) or None,
            "in_catalog": vid in catalog_ids,
        }
        if not vid:
            no_id.append(rec)
            continue
        # dedupe by youtube id: prefer the more-advanced / more-filled record
        prev = by_id.get(vid)
        if prev is None:
            by_id[vid] = rec
        else:
            if STAGE_ORDER[rec["stage"]] > STAGE_ORDER[prev["stage"]]:
                by_id[vid] = rec

    videos = sorted(by_id.values(), key=lambda x: x["title"].lower())
    videos.extend(no_id)

    # ----- aggregate stats -----
    from collections import Counter
    stage_counts = Counter(v["stage"] for v in videos)
    theme_counts = Counter(v["theme"] for v in videos if v["theme"])
    tq_counts = Counter(v["transcription"]["quality"] for v in videos)
    total_sec = sum(v["duration_sec"] or 0 for v in videos)

    payload = {
        "schema": "video_pipeline/1",
        "source": "video-archive.xlsx (sheet video_info_new), retired after migration",
        "stages": STAGES,
        "stats": {
            "videos": len(videos),
            "videos_in_catalog": sum(1 for v in videos if v["in_catalog"]),
            "total_seconds": total_sec,
            "total_hours": round(total_sec / 3600, 2),
            "by_stage": dict(stage_counts),
            "by_theme": dict(theme_counts.most_common()),
            "by_transcription_quality": dict(tq_counts),
        },
        "videos": videos,
    }
    return payload, unmapped_statuses, no_id


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true", help="parse and report, write nothing")
    args = ap.parse_args()

    payload, unmapped, no_id = build()
    s = payload["stats"]
    print(f"videos: {s['videos']}  in catalog: {s['videos_in_catalog']}  hours: {s['total_hours']}")
    print("by_stage:", json.dumps(s["by_stage"], ensure_ascii=False))
    print("by_transcription_quality:", json.dumps(s["by_transcription_quality"], ensure_ascii=False))
    if unmapped:
        print("UNMAPPED statuses (fell back to inference):")
        for u in sorted(unmapped):
            print("  ", repr(u))
    if no_id:
        print(f"rows without a YouTube id: {len(no_id)} -> {[r['title'] for r in no_id]}")

    if args.check:
        print("--check: no file written")
        return

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, sort_keys=False)
        f.write("\n")
    # BOM guard per project rule
    with open(OUT, "rb") as f:
        assert f.read(3).hex() != "efbbbf", "BOM written — fix encoding"
    print(f"wrote {os.path.relpath(OUT, ROOT)}")


if __name__ == "__main__":
    main()
